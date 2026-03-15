"""
Excel report generation (monthly transactions).
Two modes: clean export and template-based export.
"""
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.agents.admin.models import EXPORT_HEADERS, REPORTS_DIR, TEMPLATE_PATH
from app.agents.admin.repositories import (
    get_conn,
    get_all_users,
    get_budgets,
    get_expense_by_category_for_month,
)
from app.agents.admin import stats

FILL_EGRESO = PatternFill(fill_type="solid", fgColor="FFEBEE")
FILL_INGRESO = PatternFill(fill_type="solid", fgColor="E8F5E9")

CURRENCY_FORMAT = '"$"#,##0.00'


def _fix_dashboard_formulas(wb) -> None:
    """
    Corrige las fórmulas del DASHBOARD para que solo sumen egresos (type=EGRESO).
    Evita que ingresos se cuenten como gastos.
    """
    if "DASHBOARD" not in wb.sheetnames:
        return
    ws = wb["DASHBOARD"]
    # Total gastos: solo EGRESO
    ws["B3"].value = '=SUMIF(MOVIMIENTOS!E:E,"EGRESO",MOVIMIENTOS!G:G)'
    # Total Cross: solo EGRESO
    ws["B4"].value = '=SUMIFS(MOVIMIENTOS!G:G,MOVIMIENTOS!F:F,"Cross",MOVIMIENTOS!E:E,"EGRESO")'
    # Total Pau: solo EGRESO
    ws["B5"].value = '=SUMIFS(MOVIMIENTOS!G:G,MOVIMIENTOS!F:F,"Pau",MOVIMIENTOS!E:E,"EGRESO")'
    # Categorías: solo EGRESO (referencia dinámica a la celda de categoría)
    for row in range(9, 19):
        cat_cell = f"A{row}"
        ws.cell(row=row, column=2).value = (
            f'=SUMIFS(MOVIMIENTOS!G:G,MOVIMIENTOS!C:C,{cat_cell},MOVIMIENTOS!E:E,"EGRESO")'
        )


def _current_month() -> str:
    return datetime.now().strftime("%Y-%m")


def _get_excel_summary_data(month_prefix: str) -> dict:
    """
    Datos para la hoja RESUMEN: mes, ingresos por usuario, gastos, balance,
    top 5 categorías, presupuestos agregados (combinados de todos los usuarios).
    Reutiliza stats.compute_stats_all, stats.compute_stats y repositories.
    """
    st = stats.compute_stats_all(month_prefix)
    start, end = stats.month_range(month_prefix)

    conn = get_conn()
    income_by_user = []
    try:
        for user_id, user_name in get_all_users(conn):
            st_user = stats.compute_stats(user_name, month_prefix)
            income_by_user.append((user_name, st_user["income"]))
    finally:
        conn.close()

    # Gastado por categoría: suma de EGRESO de TODOS los usuarios (mismas categorías que en MOVIMIENTOS)
    all_expenses_by_cat: dict[str, float] = {}
    conn = get_conn()
    try:
        for user_id, user_name in get_all_users(conn):
            expenses = get_expense_by_category_for_month(conn, user_id, start, end)
            for cat, amt in expenses.items():
                all_expenses_by_cat[cat] = all_expenses_by_cat.get(cat, 0.0) + amt
    finally:
        conn.close()

    # Presupuestos agregados por categoría (suma de todos los usuarios)
    agg_budget: dict[str, dict] = {}
    conn = get_conn()
    try:
        for user_id, user_name in get_all_users(conn):
            rows = get_budgets(conn, user_id)
            for cat, budget_amt, curr in rows:
                budget_amt = float(budget_amt)
                if budget_amt <= 0:
                    continue
                if cat not in agg_budget:
                    agg_budget[cat] = {"budget": 0.0, "spent": 0.0, "currency": curr or "CAD"}
                agg_budget[cat]["budget"] += budget_amt
                agg_budget[cat]["spent"] = all_expenses_by_cat.get(cat, 0.0)
    finally:
        conn.close()

    budget_rows = []
    for cat, v in sorted(agg_budget.items()):
        budget = v["budget"]
        spent = v["spent"]
        available = budget - spent
        pct = (spent / budget * 100.0) if budget > 0 else 0.0
        budget_rows.append({
            "category": cat,
            "budget": budget,
            "spent": spent,
            "available": available,
            "pct": round(pct, 1),
            "currency": v["currency"],
        })

    return {
        "month": month_prefix,
        "income_by_user": income_by_user,
        "income_total": st["income"],
        "expense": st["expense"],
        "balance": st["balance"],
        "top_categories": list(st["by_cat"].items())[:5],
        "budget_rows": budget_rows,
    }


def _build_resumen_sheet(wb, month_prefix: str) -> None:
    """
    Crea la hoja RESUMEN al inicio del workbook.
    Incluye: mes, ingresos, gastos, balance, top 5 categorías, presupuestos.
    """
    data = _get_excel_summary_data(month_prefix)

    if "RESUMEN" in wb.sheetnames:
        ws = wb["RESUMEN"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("RESUMEN", 0)

    row = 1
    ws.cell(row=row, column=1).value = f"Resumen ejecutivo | {data['month']}"
    row += 2

    ws.cell(row=row, column=1).value = "Mes"
    ws.cell(row=row, column=2).value = data["month"]
    row += 1
    for user_name, inc in data["income_by_user"]:
        ws.cell(row=row, column=1).value = f"Ingresos {user_name}"
        ws.cell(row=row, column=2).value = inc
        ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
        row += 1
    ws.cell(row=row, column=1).value = "Ingresos Totales"
    ws.cell(row=row, column=2).value = data["income_total"]
    ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
    row += 1
    ws.cell(row=row, column=1).value = "Gastos"
    ws.cell(row=row, column=2).value = data["expense"]
    ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
    row += 1
    ws.cell(row=row, column=1).value = "Balance"
    ws.cell(row=row, column=2).value = data["balance"]
    ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
    row += 2

    ws.cell(row=row, column=1).value = "Top 5 categorías de gasto"
    row += 1
    if data["top_categories"]:
        ws.cell(row=row, column=1).value = "Categoría"
        ws.cell(row=row, column=2).value = "Monto"
        row += 1
        for cat, amt in data["top_categories"]:
            ws.cell(row=row, column=1).value = cat
            ws.cell(row=row, column=2).value = amt
            ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
            row += 1
    else:
        ws.cell(row=row, column=1).value = "Sin gastos en el mes"
        row += 1
    row += 1

    if data["budget_rows"]:
        ws.cell(row=row, column=1).value = "Presupuestos (combinados)"
        row += 1
        ws.cell(row=row, column=1).value = "Categoría"
        ws.cell(row=row, column=2).value = "Presupuesto"
        ws.cell(row=row, column=3).value = "Gastado"
        ws.cell(row=row, column=4).value = "Disponible"
        ws.cell(row=row, column=5).value = "% usado"
        row += 1
        for r in data["budget_rows"]:
            ws.cell(row=row, column=1).value = r["category"]
            ws.cell(row=row, column=2).value = r["budget"]
            ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
            ws.cell(row=row, column=3).value = r["spent"]
            ws.cell(row=row, column=3).number_format = CURRENCY_FORMAT
            ws.cell(row=row, column=4).value = r["available"]
            ws.cell(row=row, column=4).number_format = CURRENCY_FORMAT
            ws.cell(row=row, column=5).value = f"{r['pct']}%"
            row += 1
    else:
        ws.cell(row=row, column=1).value = "Sin presupuestos definidos"


def export_movimientos_excel(month_yyyy_mm: str | None = None) -> Path:
    """
    Export limpio (sin plantilla) para que Excel NUNCA pida reparar.
    Devuelve el Path del archivo generado.
    """
    month_prefix = month_yyyy_mm or _current_month()
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            t.date,
            t.description,
            t.category,
            t.payment_method,
            t.type,
            u.name,
            t.amount
        FROM transactions t
        JOIN users u ON u.id = t.user_id
        WHERE t.ts LIKE ?
        ORDER BY t.date ASC, t.id ASC
        """,
        (f"{month_prefix}%",),
    )
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "MOVIMIENTOS"
    ws.append(EXPORT_HEADERS)
    for row in rows:
        ws.append(list(row))

    for r, row in enumerate(rows, start=2):
        tx_type = (row[4] or "").strip().upper()
        fill = FILL_INGRESO if tx_type == "INGRESO" else (FILL_EGRESO if tx_type == "EGRESO" else None)
        if fill:
            for c in range(1, len(EXPORT_HEADERS) + 1):
                ws.cell(row=r, column=c).fill = fill

    _build_resumen_sheet(wb, month_prefix)

    out_path = REPORTS_DIR / f"movimientos_{month_prefix}.xlsx"
    wb.save(out_path)
    print(f"✅ Movimientos generado (sin plantilla): {out_path}")
    return out_path


def export_excel_template_copy(month_yyyy_mm: str | None = None) -> Path | None:
    """
    Export basado en plantilla.
    (Puede hacer que Excel pida 'reparar contenido' por validaciones/gráficos.)
    Devuelve el Path del archivo o None si la plantilla no existe.
    """
    if not TEMPLATE_PATH.exists():
        print("❌ Plantilla no encontrada")
        return None

    month_prefix = month_yyyy_mm or _current_month()
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = REPORTS_DIR / f"presupuesto_{month_prefix}.xlsx"
    shutil.copyfile(TEMPLATE_PATH, out_path)

    wb = load_workbook(out_path)
    if "MOVIMIENTOS" not in wb.sheetnames:
        ws = wb.create_sheet("MOVIMIENTOS")
    else:
        ws = wb["MOVIMIENTOS"]

    for i, h in enumerate(EXPORT_HEADERS, 1):
        ws.cell(row=1, column=i).value = h

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            t.date,
            t.description,
            t.category,
            t.payment_method,
            t.type,
            u.name,
            t.amount
        FROM transactions t
        JOIN users u ON u.id = t.user_id
        WHERE t.ts LIKE ?
        ORDER BY t.date ASC, t.id ASC
        """,
        (f"{month_prefix}%",),
    )
    rows = cur.fetchall()
    conn.close()

    r = 2
    for row in rows:
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c).value = val
        tx_type = (row[4] or "").strip().upper()
        fill = FILL_INGRESO if tx_type == "INGRESO" else (FILL_EGRESO if tx_type == "EGRESO" else None)
        if fill:
            for c in range(1, len(EXPORT_HEADERS) + 1):
                ws.cell(row=r, column=c).fill = fill
        r += 1

    _build_resumen_sheet(wb, month_prefix)
    _fix_dashboard_formulas(wb)

    wb.save(out_path)
    print(f"✅ Excel generado (plantilla): {out_path}")
    return out_path
